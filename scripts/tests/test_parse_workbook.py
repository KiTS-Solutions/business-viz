import openpyxl
from pricing_pipeline.parse_pricing import parse_workbook


def _build_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts",
               "Joe & the Juice", "Starbucks ", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None, None, None])
    ws.append(["Double Espresso Macchiato", 300000, "-", "-", "-", "-", 300000, 0])
    ws.append(["Americano MEDIUM", 350000, 400000, "-", 358000, 350000, 364500, -14500])
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return str(path)


def _config():
    return {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts", "Joe & the Juice", "Starbucks"],
    }


def test_parse_workbook_skips_dashes_and_assigns_category(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    macchiato = [r for r in result["records"] if r["product"] == "Double Espresso Macchiato"]
    assert len(macchiato) == 1
    assert macchiato[0]["brand"] == "Stories"
    assert macchiato[0]["price_lbp"] == 300000
    assert macchiato[0]["category"] == "Black Coffee"
    assert macchiato[0]["price_usd"] == 3.35


def test_parse_workbook_multiple_brands_per_product(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    americano = [r for r in result["records"] if r["product"] == "Americano MEDIUM"]
    assert len(americano) == 4
    assert {r["brand"] for r in americano} == {"Stories", "Espresso Lab", "Joe & the Juice", "Starbucks"}


def test_parse_workbook_meta_fields(tmp_path):
    xlsx_path = _build_workbook(tmp_path)
    result = parse_workbook(xlsx_path, _config())

    assert result["meta"]["client"] == "Stories"
    assert result["meta"]["own_brand"] == "Stories"
    assert result["meta"]["generated_from"] == xlsx_path


def test_parse_workbook_different_competitor_count(tmp_path):
    """Test that brand column count is derived from config, not hardcoded to 5."""
    wb = openpyxl.Workbook()
    ws = wb.active
    # 2 competitors = 3 brand columns total (1 own + 2 competitors)
    ws.append(["Products Competitors", "Stories ", "Espresso Lab", "Dunkin Donuts", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None])
    ws.append(["Americano MEDIUM", 350000, 400000, 380000, 376666, -26666])
    path = tmp_path / "sample_2competitors.xlsx"
    wb.save(path)

    config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts"],  # Only 2 competitors
    }

    result = parse_workbook(str(path), config)
    americano = [r for r in result["records"] if r["product"] == "Americano MEDIUM"]
    assert len(americano) == 3
    assert {r["brand"] for r in americano} == {"Stories", "Espresso Lab", "Dunkin Donuts"}


def test_parse_workbook_config_mismatch_raises_error(tmp_path):
    """Test that mismatched config and header brands raise ValueError."""
    xlsx_path = _build_workbook(tmp_path)

    # Config with a competitor not in the header
    bad_config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["Espresso Lab", "Dunkin Donuts", "Unknown Brand", "Starbucks"],
    }

    try:
        parse_workbook(xlsx_path, bad_config)
        assert False, "Should have raised ValueError"
    except ValueError as e:
        assert "Unknown Brand" in str(e)
        assert "do not match config's expected brands" in str(e)


def test_parse_workbook_header_has_extra_brands_not_in_config(tmp_path):
    """
    Test that raises ValueError when header has MORE brands than config expects.
    This reproduces the scenario where someone adds a 5th competitor to the
    spreadsheet (e.g., "E") but forgets to update the config.
    Before the fix, this would silently drop the extra column with no error.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    # Header has 5 brands (Stories, A, B, C, D, E) + Average + Difference
    ws.append(["Products Competitors", "Stories", "A", "B", "C", "D", "E", "Average", "Difference"])
    ws.append(["Black Coffee", None, None, None, None, None, None, None, None])
    ws.append(["Americano MEDIUM", 350000, 400000, 380000, 360000, 370000, 390000, 374333, 24333])
    path = tmp_path / "sample_extra_brands.xlsx"
    wb.save(path)

    # Config expects only 4 competitors (A, B, C, D) — missing E
    config = {
        "client": "Stories",
        "report_date": "2026-03-01",
        "currency": "LBP",
        "fx_usd_rate": 89600,
        "fx_rate_date": "2026-07-20",
        "fx_source": "test",
        "own_brand": "Stories",
        "competitors": ["A", "B", "C", "D"],  # Only 4, missing E
    }

    try:
        parse_workbook(str(path), config)
        assert False, "Should have raised ValueError when header has extra brands not in config"
    except ValueError as e:
        error_msg = str(e)
        assert "do not match config's expected brands" in error_msg
        assert "E" in error_msg
        assert "Present in header but not in config" in error_msg
