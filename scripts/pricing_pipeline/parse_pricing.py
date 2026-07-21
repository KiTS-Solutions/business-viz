def is_category_header_row(row: tuple) -> bool:
    product = row[0]
    rest = row[1:8]
    return product is not None and all(v is None for v in rest)


import openpyxl


def _clean(value):
    if isinstance(value, str):
        stripped = value.strip()
        return stripped if stripped else None
    return value


def parse_workbook(xlsx_path: str, config: dict) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    header = rows[0]

    # Locate the average column to determine the actual brand column span.
    # Most sheets label it "Average"; the Non-Dairy comparison file labels
    # it "Avg" instead — accept both rather than adding a config key for
    # what is a structural header-label variant, not client-specific data.
    average_col_index = None
    for i, col_header in enumerate(header):
        if _clean(col_header) in ("Average", "Avg"):
            average_col_index = i
            break

    if average_col_index is None:
        raise ValueError(
            "Could not locate the 'Average' column in the header row — "
            "sheet layout does not match the expected template."
        )

    # Extract the FULL actual set of brand columns from the file, then
    # translate any header text an alias covers (e.g. "Esp. Lab" -> "Espresso
    # Lab") to the canonical brand name before validating against config.
    brand_aliases = {_clean(k): _clean(v) for k, v in config.get("brand_aliases", {}).items()}
    raw_brand_columns = [_clean(h) for h in header[1:average_col_index]]
    brand_columns = [brand_aliases.get(b, b) for b in raw_brand_columns]

    # Validate that header brands match config brands (bidirectional check)
    expected_brands = {_clean(config["own_brand"]), *[_clean(c) for c in config["competitors"]]}
    actual_brands = set(brand_columns)

    if actual_brands != expected_brands:
        missing = expected_brands - actual_brands
        extra = actual_brands - expected_brands
        raise ValueError(
            f"Header brand columns {brand_columns} do not match config's expected brands "
            f"(own_brand + competitors). Missing from header: {sorted(missing) or 'none'}. "
            f"Present in header but not in config: {sorted(extra) or 'none'}."
        )

    dropped_categories = {_clean(c) for c in config.get("dropped_categories", [])}

    fx_rate = config["fx_usd_rate"]
    records = []
    current_category = None

    for row in rows[1:]:
        product = _clean(row[0])
        if product is None:
            continue

        if is_category_header_row(row):
            current_category = product
            continue

        if current_category in dropped_categories:
            continue

        for i, brand in enumerate(brand_columns):
            price = row[1 + i]
            if not isinstance(price, (int, float)):
                continue
            records.append({
                "category": current_category,
                "product": product,
                "brand": brand,
                "price_lbp": price,
                "price_usd": round(price / fx_rate, 2),
            })

    meta = {
        "client": config["client"],
        "report_date": config["report_date"],
        "currency": config["currency"],
        "fx_usd_rate": fx_rate,
        "fx_rate_date": config["fx_rate_date"],
        "fx_source": config["fx_source"],
        "own_brand": _clean(config["own_brand"]),
        "competitors": [_clean(c) for c in config["competitors"]],
        "generated_from": xlsx_path,
    }

    return {"meta": meta, "records": records}


import argparse
import json
import os

from pricing_pipeline.config import load_source_config


def run_pipeline(xlsx_path: str, config_path: str, output_path: str) -> dict:
    config = load_source_config(config_path)
    result = parse_workbook(xlsx_path, config)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    return result


def _build_arg_parser():
    parser = argparse.ArgumentParser(description="Parse a raw pricing-comparison Excel file into normalized JSON.")
    parser.add_argument("--xlsx", required=True, help="Path to the raw Excel file")
    parser.add_argument("--config", required=True, help="Path to the source config JSON")
    parser.add_argument("--out", required=True, help="Path to write the normalized JSON")
    return parser


def main(argv=None):
    args = _build_arg_parser().parse_args(argv)
    run_pipeline(args.xlsx, args.config, args.out)


if __name__ == "__main__":
    main()
